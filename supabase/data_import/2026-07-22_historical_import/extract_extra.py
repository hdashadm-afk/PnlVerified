import openpyxl
import glob
import os
import json
from extract import period_from_filename, STATION_MAP, find_label_row


def find_label_row_col1(ws, label, start=1, end=40):
    for r in range(start, end + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


def extract_manpower(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'daily cost' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['daily cost']
    header_row = 9
    stations_row = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    r_payroll = find_label_row_col1(ws, 'Total payroll')
    r_benefits = find_label_row_col1(ws, 'Total benefits')
    out = []
    if r_payroll is None or r_benefits is None:
        wb.close()
        return []
    for col, name in enumerate(stations_row, start=1):
        if name not in STATION_MAP:
            continue
        payroll = ws.cell(row=r_payroll, column=col).value or 0
        benefits = ws.cell(row=r_benefits, column=col).value or 0
        out.append({
            'station': STATION_MAP[name],
            'manpower_cost': float(payroll) + float(benefits),
        })
    wb.close()
    return out


def extract_calibration_new(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for excel_name, code in STATION_MAP.items():
        sheet_name = f'SSM-{excel_name}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        r_calib = find_label_row(ws, 'Calibration Adjustment')
        if r_calib is None:
            continue
        calib = ws.cell(row=r_calib, column=2).value or 0
        out.append({'station': code, 'calibration': float(calib)})
    wb.close()
    return out


def extract_calibration_old(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for excel_name, code in STATION_MAP.items():
        sheet_name = f'SSM-{excel_name}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        r = find_label_row_col1(ws, 'Calibration savings', start=115, end=145)
        if r is None:
            continue
        calib = ws.cell(row=r, column=2).value or 0
        out.append({'station': code, 'calibration': float(calib)})
    wb.close()
    return out


def main():
    old_format_files = sorted(glob.glob('y2024/01_*Profitability.xlsx'))
    new_format_files = sorted(glob.glob('y2024/*Profitability.xlsx')) + \
        sorted(glob.glob('y2025/*Profitability.xlsx')) + \
        sorted(glob.glob('y2026/*Profitability.xlsx'))
    new_format_files = [f for f in new_format_files if not os.path.basename(f).startswith('Format') and f not in old_format_files]

    extra = {}  # (station, period) -> {manpower_cost, calibration}

    for f in old_format_files:
        period = period_from_filename(f)
        for r in extract_manpower(f):
            extra.setdefault((r['station'], period), {})['manpower_cost'] = r['manpower_cost']
        for r in extract_calibration_old(f):
            extra.setdefault((r['station'], period), {})['calibration'] = r['calibration']
        print(f, 'old format extra done')

    for f in new_format_files:
        period = period_from_filename(f)
        for r in extract_manpower(f):
            extra.setdefault((r['station'], period), {})['manpower_cost'] = r['manpower_cost']
        for r in extract_calibration_new(f):
            extra.setdefault((r['station'], period), {})['calibration'] = r['calibration']
        print(f, 'new format extra done')

    data = json.load(open('extracted.json'))
    missing_manpower = 0
    missing_calib = 0
    for row in data:
        key = (row['station'], row['period'])
        e = extra.get(key, {})
        if 'manpower_cost' in e:
            row['manpower_cost'] = e['manpower_cost']
        else:
            row['manpower_cost'] = None
            missing_manpower += 1
        if 'calibration' in e:
            row['calibration'] = e['calibration']
        else:
            row['calibration'] = None
            missing_calib += 1

    with open('extracted_with_extra.json', 'w') as fh:
        json.dump(data, fh, indent=2)
    print(f'\nRows missing manpower_cost: {missing_manpower}')
    print(f'Rows missing calibration: {missing_calib}')
    print(f'Total rows: {len(data)}')


if __name__ == '__main__':
    main()
