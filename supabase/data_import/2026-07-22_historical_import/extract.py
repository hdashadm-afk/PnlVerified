import openpyxl
import glob
import os
import re
import json

STATION_MAP = {
    'Bani': 'BANI', 'HD': 'HD', 'HB': 'HB', 'HT': 'HT',
    'HSJ': 'HSJ', 'HQ': 'HQ', 'HM': 'HM', 'HC': 'HC',
}

MONTH_NUM = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sept': 9, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def period_from_filename(path):
    base = os.path.basename(path)
    m = re.match(r'^\d+_([A-Za-z]+)(\d{4})_', base)
    if not m:
        return None
    mon = MONTH_NUM[m.group(1).lower()]
    year = int(m.group(2))
    return f'{year:04d}-{mon:02d}-01'


def extract_old_format(path):
    # Nov2023, Dec2023, Jan2024 — separate "Profitability" sheet.
    #
    # NOTE: the Profitability sheet's row10 "Calibration savings" formula
    # is a direct cross-sheet reference into each station's own SSM-<code>
    # sheet — correct for 6/8 stations (='SSM-<code>'!$B$132, the real
    # "Calibration savings" total row) but wrong for HC and HM, whose
    # formulas instead point at $B$129 / SUM($F$129:$BD$129), which is the
    # "End Inventory - ULG" row, not calibration savings at all. For HM
    # that row happened to be empty that month (evaluates to 0, so no
    # visible damage), but for HC it summed 31 days of End Inventory ULG
    # figures and inflated "Net income" ~100x (row17 chains through
    # row11=row7+row10 and row15=row11-row14). This is a bug in the
    # founder's own source workbook, not in our extraction — verified by
    # reading both the formula and the correct calibration-savings value
    # directly from each SSM-<code> sheet's row132, which is consistently
    # labeled "Calibration savings" for all 8 stations. We recompute
    # net_profit from that ground truth rather than trusting the
    # Profitability sheet's row17 cell for stations where it disagrees.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Profitability']
    stations_row = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    out = []
    for col, name in enumerate(stations_row, start=1):
        if name not in STATION_MAP:
            continue
        # "Current MTD" is this column; "Est Full Month" is col+1
        mtd_label = ws.cell(row=4, column=col).value
        if mtd_label != 'Current MTD':
            continue
        revenue = ws.cell(row=5, column=col).value or 0
        cogs = ws.cell(row=6, column=col).value or 0
        gross_margin = ws.cell(row=7, column=col).value or 0
        opex = ws.cell(row=14, column=col).value or 0
        admin = ws.cell(row=16, column=col).value or 0

        sheet_name = f'SSM-{name}'
        if sheet_name in wb.sheetnames:
            calib_label = wb[sheet_name].cell(row=132, column=1).value
            calib = wb[sheet_name].cell(row=132, column=2).value or 0
            assert calib_label == 'Calibration savings', \
                f'{path}: {sheet_name} row132 label changed to {calib_label!r}'
        else:
            calib = ws.cell(row=10, column=col).value or 0

        net_income = float(gross_margin) + float(calib) - float(opex) - float(admin)

        out.append({
            'station': STATION_MAP[name],
            'revenue': float(revenue),
            'cogs': float(cogs),
            'opex_cost': float(opex) + float(admin),
            'net_profit': net_income,
        })
    wb.close()
    return out


def find_label_row(ws, label, start=60, end=105):
    # Row numbers drift by a row or two in some months (e.g. Oct2024's
    # SSM-HD has an extra inserted row). Locate by label text within a
    # window around the expected position instead of trusting fixed row
    # numbers everywhere.
    for r in range(start, end + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


def extract_new_format(path):
    # Feb2024 onward — per-station SSM-<code> sheet rollup
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for excel_name, code in STATION_MAP.items():
        sheet_name = f'SSM-{excel_name}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        r_sales = find_label_row(ws, 'Total Sales')
        r_margin = find_label_row(ws, 'Total margin')
        r_opex_admin = find_label_row(ws, 'Total Opex and Admin Cost')
        r_calib = find_label_row(ws, 'Est. Inc. Full Month w/ Calib')
        r_est = find_label_row(ws, 'Est. Inc. Full Month')

        if r_sales is None or r_margin is None or r_opex_admin is None:
            # structure drifted beyond the search window — skip rather
            # than silently record a wrong number
            continue

        total_sales = ws.cell(row=r_sales, column=2).value or 0
        total_margin = ws.cell(row=r_margin, column=2).value or 0
        total_opex_admin = ws.cell(row=r_opex_admin, column=2).value or 0
        net_calib = ws.cell(row=r_calib, column=2).value if r_calib else None
        if net_calib is None:
            net_calib = ws.cell(row=r_est, column=2).value or 0 if r_est else 0
        revenue = float(total_sales)
        cogs = revenue - float(total_margin)
        out.append({
            'station': code,
            'revenue': revenue,
            'cogs': cogs,
            'opex_cost': float(total_opex_admin),
            'net_profit': float(net_calib),
        })
    wb.close()
    return out


def main():
    results = []
    # Founder's decision (2026-07-22): data starts Jan 2024. Nov/Dec 2023
    # are old-format too but out of scope, and Nov2023's SSM-HM sheet has
    # a drifted row layout (row132 isn't "Calibration savings" there),
    # so there's no reason to fight that structure for months we're not
    # importing anyway.
    old_format_files = sorted(glob.glob('y2024/01_*Profitability.xlsx'))
    new_format_files = sorted(glob.glob('y2024/*Profitability.xlsx')) + \
        sorted(glob.glob('y2025/*Profitability.xlsx')) + \
        sorted(glob.glob('y2026/*Profitability.xlsx'))
    new_format_files = [f for f in new_format_files if not os.path.basename(f).startswith('Format') and f not in old_format_files]

    for f in old_format_files:
        period = period_from_filename(f)
        if not period:
            continue
        rows = extract_old_format(f)
        for r in rows:
            r['period'] = period
            r['source_file'] = os.path.basename(f)
            results.append(r)
        print(f'{f}: {len(rows)} stations (old format)')

    for f in new_format_files:
        period = period_from_filename(f)
        if not period:
            continue
        rows = extract_new_format(f)
        for r in rows:
            r['period'] = period
            r['source_file'] = os.path.basename(f)
            results.append(r)
        print(f'{f}: {len(rows)} stations (new format)')

    with open('extracted.json', 'w') as fh:
        json.dump(results, fh, indent=2)
    print(f'\nTotal rows extracted: {len(results)}')


if __name__ == '__main__':
    main()
